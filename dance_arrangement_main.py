import os
import sys

def setup_ffmpeg():
    """自动设置 FFmpeg 环境变量，确保被调用的脚本能找到 ffmpeg"""
    # 获取程序所在目录（兼容打包后和开发时）
    if getattr(sys, 'frozen', False):
        # 打包后的 .exe 所在目录
        base_dir = os.path.dirname(sys.executable)
    else:
        # 开发时 .py 文件所在目录
        base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 检查同目录下是否有 ffmpeg 文件夹
    ffmpeg_bin = os.path.join(base_dir, 'ffmpeg', 'bin')
    ffmpeg_exe = os.path.join(ffmpeg_bin, 'ffmpeg.exe')
    
    if os.path.exists(ffmpeg_exe):
        # 设置环境变量，这样 subprocess 调用的程序也能继承
        os.environ['PATH'] = ffmpeg_bin + os.pathsep + os.environ.get('PATH', '')
        os.environ['FFMPEG_PATH'] = ffmpeg_exe
        print(f"[信息] 已加载 FFmpeg: {ffmpeg_exe}")
    else:
        # 如果没找到，给出友好提示但继续运行（可能用户手动配置了）
        print(f"[信息] 未在 {ffmpeg_bin} 找到 ffmpeg.exe，将尝试使用系统 PATH")

# ========== 自动初始化 ==========
setup_ffmpeg()

import random
import os
import glob
import warnings
import pandas as pd
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from ttkbootstrap import Style
from ttkbootstrap import Messagebox
import openpyxl
from openpyxl import load_workbook
from collections import Counter
warnings.filterwarnings('ignore')

# ---------- 全局数据 ----------
DATA = []
POOL = []
TOTAL = 0

class PlaylistApp:
    def __init__(self, root):
        self.root = root
        root.title("舞曲排布工具-魅影制作")
        
        # ---------- 窗口居中 ----------
        window_width = 1800
        window_height = 1253
        screen_width = root.winfo_screenwidth()
        screen_height = root.winfo_screenheight()
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2 - 30
        root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        root.minsize(1600, 800)

        # ---------- 全局样式设置 ----------
        style = Style(theme="pulse-dark")
        style.configure("TLabel", font=("Microsoft YaHei", 9))
        style.configure("TLabelframe", font=("Microsoft YaHei", 10, "bold"))
        style.configure("TLabelframe.Label", font=("Microsoft YaHei", 10, "bold"))
        style.configure("TButton", font=("Microsoft YaHei", 9))
        style.configure("TCheckbutton", font=("Microsoft YaHei", 9))
        style.configure("TEntry", font=("Microsoft YaHei", 9))
        style.configure("Status.TLabel", font=("Microsoft YaHei", 9), relief="sunken", padding=5)

        # ---------- 主布局：左右结构，使用grid ----------
        main_frame = ttk.Frame(root, padding=(10, 10, 10, 10))
        main_frame.pack(fill="both", expand=True)
        
        # 配置grid列权重
        main_frame.columnconfigure(0, weight=0, minsize=600)  # 左侧固定宽度
        main_frame.columnconfigure(1, weight=3)  # 右侧可扩展
        main_frame.rowconfigure(0, weight=1)

        # -------- 左侧：控制面板 --------
        left_frame = ttk.Frame(main_frame)
        left_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        left_frame.columnconfigure(0, weight=1)
        
        # ---------- 数据加载区 ----------
        frame_data = ttk.Labelframe(left_frame, text=" 数据源 ", padding=(10, 8), bootstyle="primary")
        frame_data.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        frame_data.columnconfigure(0, weight=1)
        
        # 自动查找Excel文件
        default_path = self.find_excel_file()
        self.data_path = ttk.StringVar(value=default_path)
        
        # 路径输入框
        self.path_entry = ttk.Entry(frame_data, textvariable=self.data_path, width=40)
        self.path_entry.grid(row=0, column=0, sticky="ew", padx=(0, 8), pady=5)
        
        # 按钮框架
        btn_data_frame = ttk.Frame(frame_data)
        btn_data_frame.grid(row=0, column=1, sticky="e", pady=5)
        
        ttk.Button(btn_data_frame, text="浏览", command=self.browse_file, 
                  bootstyle="secondary", width=8).pack(side="left", padx=(0, 5))
        ttk.Button(btn_data_frame, text="加载", command=self.load_data, 
                  bootstyle="primary", width=8).pack(side="left")
        
        self.data_status = ttk.Label(frame_data, text="未加载", bootstyle="danger", 
                                     font=("Microsoft YaHei", 8))
        self.data_status.grid(row=1, column=0, columnspan=2, sticky="w", pady=(3, 0))
        
        # ---------- 固定位置设置区 ----------
        frame_fixed = ttk.Labelframe(left_frame, text=" 固定位置（位置=舞种，如 2=华尔兹） ", 
                                     padding=(10, 8), bootstyle="warning")
        frame_fixed.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        frame_fixed.columnconfigure(0, weight=1)
        frame_fixed.columnconfigure(1, weight=1)
        frame_fixed.columnconfigure(2, weight=1)
        
        self.fixed_positions = {}
        self.fixed_entries = []
        
        for col_idx in range(3):
            cell = ttk.Frame(frame_fixed, padding=(5, 3), relief="ridge", borderwidth=1)
            cell.grid(row=0, column=col_idx, padx=8, pady=3, sticky="ew")
            
            pos_entry = ttk.Entry(cell, width=5)
            pos_entry.pack(side="left", padx=2)
            ttk.Label(cell, text="=", font=("Microsoft YaHei", 9)).pack(side="left")
            dance_entry = ttk.Entry(cell, width=14)
            dance_entry.pack(side="left", padx=2, fill="x", expand=True)
            
            self.fixed_entries.append((pos_entry, dance_entry))
        
        btn_frame = ttk.Frame(frame_fixed)
        btn_frame.grid(row=1, column=0, columnspan=3, pady=(8, 0))
        
        ttk.Button(btn_frame, text="应用固定位置", command=self.apply_fixed_positions, 
                  bootstyle="warning", width=14).pack(side="left", padx=(0, 10))
        ttk.Button(btn_frame, text="清空", command=self.clear_fixed_positions, 
                  bootstyle="danger", width=8).pack(side="left")
        
        # ---------- 规则勾选区 ----------
        frame_rules = ttk.Labelframe(left_frame, text=" 规则列表 ", padding=(10, 8), bootstyle="info")
        frame_rules.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        frame_rules.columnconfigure(0, weight=1)
        
        self.rules = {
            '同舞种间隔': ttk.BooleanVar(value=True),
            '大类不连续三次': ttk.BooleanVar(value=True),
            '节奏约束': ttk.BooleanVar(value=True),
        }
        
        self.gap_value = ttk.StringVar(value="6")
        
        # 规则1：同舞种间隔
        row1 = ttk.Frame(frame_rules)
        row1.grid(row=0, column=0, sticky="w", pady=2)
        ttk.Checkbutton(row1, variable=self.rules['同舞种间隔'], 
                       bootstyle="round-toggle").pack(side="left", padx=(0, 5))
        ttk.Label(row1, text="同一舞种至少隔开").pack(side="left")
        gap_entry = ttk.Entry(row1, width=5, textvariable=self.gap_value)
        gap_entry.pack(side="left", padx=5)
        ttk.Label(row1, text="首（位置差≥N）").pack(side="left")
        
        # 规则2：大类不连续三次
        row2 = ttk.Frame(frame_rules)
        row2.grid(row=1, column=0, sticky="w", pady=2)
        ttk.Checkbutton(row2, variable=self.rules['大类不连续三次'], 
                       bootstyle="round-toggle").pack(side="left", padx=(0, 5))
        ttk.Label(row2, text="连续3首大类不能相同").pack(side="left")
        
        # 规则3：节奏约束
        row3 = ttk.Frame(frame_rules)
        row3.grid(row=2, column=0, sticky="w", pady=2)
        ttk.Checkbutton(row3, variable=self.rules['节奏约束'], 
                       bootstyle="round-toggle").pack(side="left", padx=(0, 5))
        ttk.Label(row3, text="中速/慢速不三连，不允许快-快").pack(side="left")
        
        # ---------- 搜索参数、控制按钮和状态信息 ----------
        frame_params = ttk.Labelframe(left_frame, text=" 搜索控制 ", padding=(10, 8), bootstyle="neutral")
        frame_params.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        frame_params.columnconfigure(0, weight=1)
        frame_params.columnconfigure(1, weight=1)

        # 搜索次数（第一行）
        self.max_attempts = ttk.StringVar(value="50000")
        entry = ttk.Entry(frame_params, width=8, textvariable=self.max_attempts, justify='center')
        entry.grid(row=0, column=0, columnspan=2, sticky="", padx=20, pady=5)

        # 控制按钮（第二行）
        frame_btn = ttk.Frame(frame_params)
        frame_btn.grid(row=1, column=0, columnspan=2, pady=(10, 5))

        self.btn_generate = ttk.Button(frame_btn, text="开始搜索", command=self.run_search,
                                    bootstyle="success", width=14)
        self.btn_generate.pack(side="left", padx=(0, 10))

        self.btn_stop = ttk.Button(frame_btn, text="停止搜索", command=self.stop_search,
                                bootstyle="danger", width=14)
        self.btn_stop.pack(side="left")
        self.btn_stop.config(state="disabled")

        # 状态信息条（第三行）
        self.progress_label = ttk.Label(frame_params, text="（注意：舞种排比和规则设定都要在合理范围内！）等待开始...", 
                                        bootstyle="secondary", style="Status.TLabel")
        self.progress_label.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(5, 0))

        # ========== 功能按钮区域 ==========
        frame_func_btns = ttk.Labelframe(left_frame, text=" 功能模块 ", padding=(10, 8), bootstyle="secondary")
        frame_func_btns.grid(row=4, column=0, sticky="ew", pady=(0, 5))
        frame_func_btns.columnconfigure(0, weight=1)
        frame_func_btns.columnconfigure(1, weight=2)
        frame_func_btns.columnconfigure(2, weight=1)

        # 标题行
        ttk.Label(frame_func_btns, text="代号", font=("Microsoft YaHei", 10, "bold"), 
                 width=6, anchor="center").grid(row=0, column=0, padx=2, pady=2)
        ttk.Label(frame_func_btns, text="配置", font=("Microsoft YaHei", 10, "bold"), 
                 width=15, anchor="center").grid(row=0, column=1, padx=2, pady=2)
        ttk.Label(frame_func_btns, text="功能名称", font=("Microsoft YaHei", 10, "bold"), 
                 width=15, anchor="center").grid(row=0, column=2, padx=2, pady=2)

        # 存储配置
        self.config_values = {
            'A': ttk.StringVar(value="标题 + 艺术家 + 专辑"),
            'B': ttk.StringVar(value="320k CBR 并发2"),
            'B_bitrate': ttk.StringVar(value="320k"),
            'B_use_vbr': ttk.BooleanVar(value=False),
            'B_vbr_quality': ttk.StringVar(value="0"),
            'B_workers': ttk.StringVar(value="2"),
            'C': ttk.StringVar(value="淡入0.1s 淡出5.0s"),
            'C_fade_in': ttk.StringVar(value="0.1"),
            'C_fade_out': ttk.StringVar(value="5.0"),
            'D': ttk.StringVar(value="时长8s 音量0.2 位置middle"),
            'D_duration': ttk.StringVar(value="8"),
            'D_start': ttk.StringVar(value="0"),
            'D_position': ttk.StringVar(value="middle"),
            'D_custom_position': ttk.StringVar(value="0"),
            'D_volume': ttk.StringVar(value="0.2")
        }

        # 功能配置列表
        func_configs = [
            ('A', '标签筛选', self.func1),
            ('B', '格式转换', self.func2),
            ('C', '时长裁剪', self.func3),
            ('D', '前缀添加', self.func4)
        ]

        self.func_buttons = []
        self.config_buttons = []

        for idx, (code, name, command) in enumerate(func_configs):
            # 代号
            ttk.Label(frame_func_btns, text=code, font=("Microsoft YaHei", 10), 
                     width=6, anchor="center").grid(row=idx+1, column=0, padx=2, pady=3)
            
            # 配置按钮
            config_btn = ttk.Button(
                frame_func_btns,
                textvariable=self.config_values[code],
                width=20,
                bootstyle="secondary",
                command=lambda c=code: self.edit_config(c)
            )
            config_btn.grid(row=idx+1, column=1, padx=2, pady=3, sticky="ew")
            self.config_buttons.append(config_btn)
            
            # 功能按钮
            func_btn = ttk.Button(
                frame_func_btns,
                text=name,
                width=12,
                bootstyle="success",
                command=command
            )
            func_btn.grid(row=idx+1, column=2, padx=2, pady=3)
            self.func_buttons.append(func_btn)

        # -------- 右侧：排布结果 --------
        right_frame = ttk.Frame(main_frame)
        right_frame.grid(row=0, column=1, sticky="nsew")
        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(0, weight=1)

        frame_result = ttk.Labelframe(right_frame, text=" 排布结果 ", padding=(10, 8), bootstyle="primary")
        frame_result.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
        frame_result.columnconfigure(0, weight=1)
        frame_result.rowconfigure(0, weight=1)

        # 结果文本框
        self.result_text = ttk.ScrolledText(frame_result, height=30, font=("Courier", 9))
        self.result_text.grid(row=0, column=0, sticky="nsew")

        # ---------- 保存排布结果按钮 ----------
        button_frame = ttk.Frame(right_frame)
        button_frame.grid(row=1, column=0, pady=(0, 5))

        self.btn_save_result = ttk.Button(
            button_frame, 
            text="保存排布结果", 
            command=self.save_result_to_excel,
            bootstyle="success", 
            width=20,
            state="disabled"  # 初始禁用，找到结果后启用
        )
        self.btn_save_result.pack()

        # 状态
        self.searching = False
        self.data_loaded = False
        self.last_dance_list = None  # 存储最后一次生成的舞种列表
        self.last_seq = None  # 存储最后一次生成的完整序列
        self.last_max_attempts = 0  # 存储最后一次搜索次数
        self.last_attempt = 0  # 存储最后一次成功搜索的尝试次数

        # 检查是否需要自动加载
        if os.path.exists(self.data_path.get()):
            self.load_data(show_message=False)
        else:
            # 如果没有找到Excel，提示用户
            self.result_text.insert("end", "未找到Excel数据文件。\n")
            self.result_text.insert("end", "请点击'浏览'选择文件。\n")
    
    # ---------- 查找Excel文件 ----------
    def find_excel_file(self):
        """在同目录下查找Excel文件"""
        # 获取程序所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # 查找所有Excel文件
        excel_files = []
        for ext in ['*.xlsx', '*.xls']:
            excel_files.extend(glob.glob(os.path.join(script_dir, ext)))
        
        if excel_files:
            # 如果找到多个，返回第一个
            return excel_files[0]
        
        # 检查示例文件
        sample_path = os.path.join(script_dir, "舞曲排布_示例.xlsx")
        if os.path.exists(sample_path):
            return sample_path
        
        return ""
    
    # ---------- 浏览文件 ----------
    def browse_file(self):
        """浏览选择Excel文件"""
        file_path = ttk.filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.data_path.set(file_path)
            self.load_data()

    # ---------- 数据加载 ----------
    def load_data(self, show_message=True):
        path = self.data_path.get()
        
        # 如果路径为空或文件不存在
        if not path or not os.path.exists(path):
            if show_message:
                response = Messagebox.yesno(
                    title="文件未找到",
                    message=f"未找到数据文件。是否创建示例文件？",
                    parent=self.root 
                )
                # 修复：ttkbootstrap 的 Messagebox.yesno 返回字符串 "yes" 或 "no"
                # 需要明确判断返回值 原写法 if response:
                if response == "yes" or response == "Yes" or response is True:                
                    sample_path = self.create_sample_excel()
                    if sample_path:
                        self.data_path.set(sample_path)
                        self.load_data(show_message=False)
                        return
            else:
                self.data_status.config(text="未加载", bootstyle="danger")
            return
        
        try:
            df = pd.read_excel(path, sheet_name="舞种配比", header=None, usecols="A:D", skiprows=1)
            df = df.dropna(how='all')
            
            global DATA, POOL, TOTAL
            DATA = []
            POOL = []
            
            for _, row in df.iterrows():
                cat = str(row[0]).strip()
                dance = str(row[1]).strip()
                rhythm = str(row[2]).strip()
                try:
                    count = int(row[3])
                except:
                    count = 1
                
                if cat and dance and rhythm:
                    DATA.append((cat, dance, rhythm, count))
                    for _ in range(count):
                        POOL.append((cat, dance, rhythm))
            
            TOTAL = len(POOL)
            self.data_status.config(text=f"已加载: {len(DATA)} 种, {TOTAL} 首", bootstyle="success")
            self.data_loaded = True
            self.result_text.delete("1.0", "end")
            self.result_text.insert("end", f"✅ 数据加载成功！\n")
            self.result_text.insert("end", f"文件: {os.path.basename(path)}\n")
            self.result_text.insert("end", f"共 {len(DATA)} 个舞种，总计 {TOTAL} 首\n")
            self.result_text.insert("end", "-" * 38 + "\n")
            for cat, dance, rhythm, count in DATA:
                self.result_text.insert("end", f"{cat:<8} {dance:<12} {rhythm:<4} x{count}\n")
            
        except Exception as e:
            Messagebox.show_error(
                title="加载失败",
                message=f"文件加载失败:\n{str(e)}\n\n请确保文件格式正确。",
                parent=self.root
            )
            self.data_status.config(text="加载失败", bootstyle="danger")
    
    # ---------- 创建示例Excel ----------
    def create_sample_excel(self):
        """创建示例Excel文件"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "舞种配比"
            
            # 写入标题行
            headers = ["大类", "舞种", "节奏", "数量"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 示例数据
            sample_data = [
                ["摩登", "华尔兹", "慢", 3],
                ["摩登", "探戈", "中", 2],
                ["摩登", "狐步", "慢", 2],
                ["摩登", "快步", "快", 2],
                ["摩登", "维也纳华尔兹", "快", 2],
                ["拉丁", "伦巴", "慢", 3],
                ["拉丁", "恰恰", "中", 3],
                ["拉丁", "桑巴", "快", 2],
                ["拉丁", "斗牛", "中", 2],
                ["拉丁", "牛仔", "快", 2],
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "舞曲排布_示例简版.xlsx")
            wb.save(file_path)
            return file_path
        except Exception as e:
            Messagebox.show_error(
                title="错误",
                message=f"创建示例文件失败：{str(e)}",
                parent=self.root
            )
            return None
    
    # ---------- 固定位置管理 ----------
    def apply_fixed_positions(self):
        self.fixed_positions = {}
        for pos_entry, dance_entry in self.fixed_entries:
            pos_text = pos_entry.get().strip()
            dance_text = dance_entry.get().strip()
            if pos_text and dance_text:
                try:
                    pos = int(pos_text)
                    self.fixed_positions[pos] = dance_text
                except:
                    pass
        self.update_fixed_display()
        Messagebox.show_info(
            title="提示",  
            message=f"已应用 {len(self.fixed_positions)} 个固定位置",
            parent=self.root  
        )

    
    def clear_fixed_positions(self):
        self.fixed_positions = {}
        for pos_entry, dance_entry in self.fixed_entries:
            pos_entry.delete(0, "end")
            dance_entry.delete(0, "end")
        self.update_fixed_display()
    
    def update_fixed_display(self):
        if self.fixed_positions:
            self.data_status.config(text=f"已加载 + 固定: {self.fixed_positions}", bootstyle="info")
        else:
            self.data_status.config(text=f"已加载: {len(DATA)} 种, {TOTAL} 首", bootstyle="success")
    
    # ---------- 约束检查 ----------
    def check_constraints(self, seq, rules):
        errors = []
        
        # 固定位置检查
        for pos, dance in self.fixed_positions.items():
            if pos < 1 or pos > len(seq):
                errors.append(f"固定位置 {pos} 超出范围")
            elif seq[pos-1][1] != dance:
                errors.append(f"位置 {pos} 应为 '{dance}'，实际为 '{seq[pos-1][1]}'")
        
        if rules.get('同舞种间隔', False):
            try:
                min_gap = int(self.gap_value.get())
            except:
                min_gap = 6
            last = {}
            for i, item in enumerate(seq):
                d = item[1]
                if d in last:
                    if i - last[d] < min_gap:
                        errors.append(f"'{d}' 位置 {last[d]+1} 和 {i+1} 间隔 {i-last[d]} < {min_gap}")
                last[d] = i
        
        if rules.get('大类不连续三次', False):
            for i in range(len(seq) - 2):
                if seq[i][0] == seq[i+1][0] == seq[i+2][0]:
                    errors.append(f"位置 {i+1}-{i+3} 大类连续三次: {seq[i][0]}")
        
        if rules.get('节奏约束', False):
            for i in range(len(seq) - 1):
                if seq[i][2] == "快" and seq[i+1][2] == "快":
                    errors.append(f"位置 {i+1}-{i+2} 快-快连续")
            for i in range(len(seq) - 2):
                if seq[i][2] == "中" and seq[i+1][2] == "中" and seq[i+2][2] == "中":
                    errors.append(f"位置 {i+1}-{i+3} 中速三连")
            for i in range(len(seq) - 2):
                if seq[i][2] == "慢" and seq[i+1][2] == "慢" and seq[i+2][2] == "慢":
                    errors.append(f"位置 {i+1}-{i+3} 慢速三连")
        
        return len(errors) == 0, errors
    
    # ---------- 生成序列 ----------
    def generate_sequence(self, rules, max_attempts=1):
        if not POOL:
            return None
        
        pool = POOL[:]
        fixed_pos = self.fixed_positions
        
        for _ in range(max_attempts):
            random.shuffle(pool)
            seq = pool[:]
            
            # 处理固定位置
            ok = True
            for pos, dance in fixed_pos.items():
                if pos < 1 or pos > len(seq):
                    ok = False
                    break
                if seq[pos-1][1] != dance:
                    swapped = False
                    for i in range(len(seq)):
                        if seq[i][1] == dance and i != pos-1:
                            seq[pos-1], seq[i] = seq[i], seq[pos-1]
                            swapped = True
                            break
                    if not swapped:
                        ok = False
                        break
            if not ok:
                continue
            
            if len(seq) != TOTAL:
                continue
            
            ok, _ = self.check_constraints(seq, rules)
            if ok:
                return seq
        
        return None
    
    # ---------- 搜索 ----------
    def stop_search(self):
        self.searching = False
        self.progress_label.config(text="已停止搜索")
        self.btn_generate.config(state="normal")
        self.btn_stop.config(state="disabled")

    # ---------- 功能函数（保持不变） ----------
    def func1(self):
        """标签筛选 - A_music_processor.py"""
        try:
            import subprocess
            import os
            
            # 获取配置值
            config_value = self.config_values['A'].get()
            
            # 解析配置，生成参数
            # 例如："标题 + 艺术家" 或 "标题 + 艺术家 + 专辑"
            params = []
            if "标题" in config_value:
                params.append("--title")
            if "艺术家" in config_value:
                params.append("--artist")
            if "专辑" in config_value:
                params.append("--album")
            
            script_dir = os.path.dirname(os.path.abspath(__file__))
            py_file = os.path.join(script_dir, "A_music_processor.py")
            
            if os.path.exists(py_file):
                # 构建命令
                cmd = ["python", py_file] + params
                subprocess.Popen(cmd)
                Messagebox.show_info(
                    title="标签筛选",  
                    message=f"已启动标签筛选程序\n匹配条件：{config_value}",
                    parent=self.root  
                )

            else:
                Messagebox.show_error(
                    title="错误",
                    message=f"找不到文件：{py_file}",
                    parent=self.root
                )
        except Exception as e:
            Messagebox.show_error(
                title="错误",
                message=f"执行失败：{str(e)}",
                parent=self.root
            )
    
    def func2(self):
        """格式转换 - B_audio_converter.py"""
        try:
            import subprocess
            import os
            
            # 获取配置值
            bitrate = self.config_values['B_bitrate'].get()
            use_vbr = self.config_values['B_use_vbr'].get()
            vbr_quality = self.config_values['B_vbr_quality'].get()
            workers = self.config_values['B_workers'].get()
            
            script_dir = os.path.dirname(os.path.abspath(__file__))
            py_file = os.path.join(script_dir, "B_audio_converter.py")
            
            if os.path.exists(py_file):
                # 构建命令
                cmd = [
                    "python", py_file,
                    "--bitrate", bitrate,
                    "--vbr" if use_vbr else "--cbr",
                    "--vbr-quality", vbr_quality,
                    "--workers", workers
                ]
                subprocess.Popen(cmd)
                Messagebox.show_info(
                    title="格式转换",  
                    message=f"已启动格式转换程序\n比特率：{bitrate}  {'VBR' if use_vbr else 'CBR'}\n并发数：{workers}",
                    parent=self.root  
                )

            else:
                Messagebox.show_error(
                    title="错误",
                    message=f"找不到文件：{py_file}",
                    parent=self.root
                )
        except Exception as e:
            Messagebox.show_error(
                title="错误",
                message=f"执行失败：{str(e)}",
                parent=self.root
            )

    def func3(self):
        """时长裁剪 - C_crop_audio.py"""
        try:
            import subprocess
            import os
            
            # 获取配置值
            fade_in = self.config_values['C_fade_in'].get()
            fade_out = self.config_values['C_fade_out'].get()
            
            script_dir = os.path.dirname(os.path.abspath(__file__))
            py_file = os.path.join(script_dir, "C_crop_audio.py")
            
            if os.path.exists(py_file):
                # 构建命令
                cmd = [
                    "python", py_file,
                    "--fade-in", fade_in,
                    "--fade-out", fade_out
                ]
                subprocess.Popen(cmd)
                Messagebox.show_info(
                    title="时长裁剪",  
                    message=f"已启动时长裁剪程序\n淡入：{fade_in}s  淡出：{fade_out}s",
                    parent=self.root  
                )

            else:
                Messagebox.show_error(
                    title="错误",
                    message=f"找不到文件：{py_file}",
                    parent=self.root
                )
        except Exception as e:
            Messagebox.show_error(
                title="错误",
                message=f"执行失败：{str(e)}",
                parent=self.root
            )

    def func4(self):
        """前缀添加 - D_add_audio.py"""
        try:
            import subprocess
            import os
            
            # 获取配置值
            duration = self.config_values['D_duration'].get()
            start = self.config_values['D_start'].get()
            position = self.config_values['D_position'].get()
            custom_position = self.config_values['D_custom_position'].get()
            volume = self.config_values['D_volume'].get()
            
            # 处理位置参数
            if position == "自定义秒数":
                final_position = custom_position
            else:
                final_position = position
            
            script_dir = os.path.dirname(os.path.abspath(__file__))
            py_file = os.path.join(script_dir, "D_add_audio.py")
            
            if os.path.exists(py_file):
                # 构建命令
                cmd = [
                    "python", py_file,
                    "--duration", duration,
                    "--start", start,
                    "--position", final_position,
                    "--volume", volume
                ]
                subprocess.Popen(cmd)
                Messagebox.show_info(
                    title="前缀添加",  
                    message=f"已启动前缀添加程序\nBGM时长：{duration}s\n音量：{volume}\n语音位置：{final_position}",
                    parent=self.root  
                )

            else:
                Messagebox.show_error(
                    title="错误",
                    message=f"找不到文件：{py_file}",
                    parent=self.root
                )
        except Exception as e:
            Messagebox.show_error(
                title="错误",
                message=f"执行失败：{str(e)}",
                parent=self.root
            )

    def edit_config(self, code):
        """编辑配置 - 根据代号弹出不同的配置对话框"""
        
        if code == 'A':
            # ===== 功能A的专用配置对话框 =====
            dialog = ttk.Toplevel(self.root)
            dialog.title("标签筛选配置")
            dialog.geometry("450x400")
            
            # 弹窗居中
            root_x = self.root.winfo_rootx()
            root_y = self.root.winfo_rooty()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            x = root_x + (root_width - 450) // 2
            y = root_y + (root_height - 400) // 2
            dialog.geometry(f"450x400+{x}+{y}")
            
            dialog.transient(self.root)
            dialog.grab_set()
            
            ttk.Label(dialog, text="选择需要匹配的标签", font=("Microsoft YaHei", 12, "bold")).pack(pady=10)
            ttk.Label(dialog, text="满足勾选以下条件复选框的舞曲才能移动", font=("Microsoft YaHei", 9)).pack()
            
            # 存储复选框状态
            self.check_vars = {
                'title': ttk.BooleanVar(value=True),
                'artist': ttk.BooleanVar(value=True),
                'album': ttk.BooleanVar(value=True)
            }
            # 三个复选框
            check_frame = ttk.Frame(dialog)
            check_frame.pack(pady=15)
            
            ttk.Checkbutton(check_frame, text="标题 (Title)", variable=self.check_vars['title'], 
                           bootstyle="round-toggle").pack(anchor="w", pady=3)
            ttk.Checkbutton(check_frame, text="艺术家 (Artist)", variable=self.check_vars['artist'], 
                           bootstyle="round-toggle").pack(anchor="w", pady=3)
            ttk.Checkbutton(check_frame, text="专辑 (Album)", variable=self.check_vars['album'], 
                           bootstyle="round-toggle").pack(anchor="w", pady=3)
            
            # 显示当前配置
            current_config = self.config_values['A'].get()
            ttk.Label(dialog, text=f"当前配置：{current_config}", 
                     font=("Microsoft YaHei", 9), bootstyle="info").pack(pady=5)
            
            # 按钮
            btn_frame = ttk.Frame(dialog)
            btn_frame.pack(pady=15)
            
            def save_config():
                # 获取勾选的状态
                selected = []
                if self.check_vars['title'].get():
                    selected.append("标题")
                if self.check_vars['artist'].get():
                    selected.append("艺术家")
                if self.check_vars['album'].get():
                    selected.append("专辑")
                
                if not selected:
                    Messagebox.show_warning(
                        title="提示",
                        message="请至少选择一个条件！",
                        parent=dialog
                    )
                    return
                
                # 生成配置字符串
                new_value = " + ".join(selected)
                self.config_values['A'].set(new_value)
                dialog.destroy()
                Messagebox.show_info(
                    title="成功",  
                    message=f"配置已更新为：{new_value}",
                    parent=self.root  
                )
            
            ttk.Button(btn_frame, text="确定", command=save_config, bootstyle="success", width=10).pack(side="left", padx=10)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy, bootstyle="secondary", width=10).pack(side="left", padx=10)            

        elif code == 'B':
            # ===== 功能B的专用配置对话框 =====
            dialog = ttk.Toplevel(self.root)
            dialog.title("格式转换配置")
            dialog.geometry("650x500")
            
            # 弹窗居中
            root_x = self.root.winfo_rootx()
            root_y = self.root.winfo_rooty()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            x = root_x + (root_width - 650) // 2
            y = root_y + (root_height - 500) // 2
            dialog.geometry(f"650x500+{x}+{y}")
            
            dialog.transient(self.root)
            dialog.grab_set()
            
            ttk.Label(dialog, text="音频转换配置", font=("Microsoft YaHei", 12, "bold")).pack(pady=10)
            # 配置框架
            config_frame = ttk.Frame(dialog)
            config_frame.pack(pady=10, padx=20, fill="x")
            
            # 1. 比特率选择
            row1 = ttk.Frame(config_frame)
            row1.pack(fill="x", pady=5)
            ttk.Label(row1, text="CBR固定比特率：", width=12, anchor="w").pack(side="left")
            bitrate_var = ttk.StringVar(value=self.config_values['B_bitrate'].get())
            bitrate_combo = ttk.Combobox(row1, textvariable=bitrate_var, values=["128k", "192k", "256k", "320k"], width=10)
            bitrate_combo.pack(side="left", padx=5)
            
            # 2. VBR开关 
            row2 = ttk.Frame(config_frame)
            row2.pack(fill="x", pady=5)
            ttk.Label(row2, text="VBR变比特率：", width=12, anchor="w").pack(side="left")
            use_vbr_var = ttk.BooleanVar(value=self.config_values['B_use_vbr'].get())
            ttk.Checkbutton(row2, variable=use_vbr_var, bootstyle="round-toggle").pack(side="left")
            ttk.Label(row2, text="（勾选后使用VBR质量）", font=("Microsoft YaHei", 8)).pack(side="left", padx=5)
            
            # 3. VBR质量（仅当VBR启用时可用）
            row3 = ttk.Frame(config_frame)
            row3.pack(fill="x", pady=5)
            ttk.Label(row3, text="VBR质量：", width=12, anchor="w").pack(side="left")
            vbr_quality_var = ttk.StringVar(value=str(self.config_values['B_vbr_quality'].get()))
            vbr_quality_combo = ttk.Combobox(row3, textvariable=vbr_quality_var, values=["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"], width=10)
            vbr_quality_combo.pack(side="left", padx=5)
            ttk.Label(row3, text="（0最高，9最低）", font=("Microsoft YaHei", 8)).pack(side="left", padx=5)

            # 4. 并发数
            row4 = ttk.Frame(config_frame)
            row4.pack(fill="x", pady=5)
            ttk.Label(row4, text="并发数：", width=12, anchor="w").pack(side="left")
            workers_var = ttk.StringVar(value=str(self.config_values['B_workers'].get()))
            ttk.Spinbox(row4, from_=1, to=8, textvariable=workers_var, width=10).pack(side="left", padx=5)
            ttk.Label(row4, text="（同时转换文件数）", font=("Microsoft YaHei", 8)).pack(side="left", padx=5)
            # 显示当前配置
            current_config = self.config_values['B'].get()
            ttk.Label(dialog, text=f"当前配置：{current_config}", 
                     font=("Microsoft YaHei", 9), bootstyle="info").pack(pady=5)

            # 按钮
            btn_frame = ttk.Frame(dialog)
            btn_frame.pack(pady=15)
            
            def save_config_b():
                bitrate = bitrate_var.get()
                use_vbr = use_vbr_var.get()
                vbr_quality = vbr_quality_var.get()
                workers = workers_var.get()

                # 验证
                try:
                    int(workers)
                except:
                    Messagebox.show_warning(
                        title="提示",
                        message="并发数请输入有效数字！",
                        parent=dialog
                    )
                    return
                
                if bitrate not in ["128k", "192k", "256k", "320k"]:
                    Messagebox.show_warning(
                        title="提示",
                        message="请选择有效的比特率！",
                        parent=dialog
                    )
                    return

                # 保存到各自的变量
                self.config_values['B_bitrate'] = ttk.StringVar(value=bitrate)
                self.config_values['B_use_vbr'] = ttk.BooleanVar(value=use_vbr)
                self.config_values['B_vbr_quality'] = ttk.StringVar(value=vbr_quality)
                self.config_values['B_workers'] = ttk.StringVar(value=workers)

                # 生成显示文本
                vbr_text = "VBR" if use_vbr else "CBR"
                display_text = f"{bitrate} {vbr_text} 并发{workers}"
                self.config_values['B'].set(display_text)
                
                dialog.destroy()
                Messagebox.show_info(
                    title="成功",  
                    message=f"配置已更新为：{display_text}",
                    parent=self.root  
                )
            
            ttk.Button(btn_frame, text="确定", command=save_config_b, bootstyle="success", width=10).pack(side="left", padx=10)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy, bootstyle="secondary", width=10).pack(side="left", padx=10)

        elif code == 'C':
            # ===== 功能C的专用配置对话框 =====
            dialog = ttk.Toplevel(self.root)
            dialog.title("时长裁剪配置")
            dialog.geometry("420x450")

            # 弹窗居中
            root_x = self.root.winfo_rootx()
            root_y = self.root.winfo_rooty()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            x = root_x + (root_width - 420) // 2
            y = root_y + (root_height - 450) // 2
            dialog.geometry(f"420x450+{x}+{y}")
            
            dialog.transient(self.root)
            dialog.grab_set()
            
            ttk.Label(dialog, text="淡入淡出配置", font=("Microsoft YaHei", 12, "bold")).pack(pady=10)

            # 配置框架
            config_frame = ttk.Frame(dialog)
            config_frame.pack(pady=10, padx=20, fill="x")

            # 1. 淡入时长
            row1 = ttk.Frame(config_frame)
            row1.pack(fill="x", pady=8)
            ttk.Label(row1, text="淡入时长（秒）：", width=14, anchor="w").pack(side="left")
            fade_in_var = ttk.StringVar(value=self.config_values['C_fade_in'].get())
            ttk.Entry(row1, textvariable=fade_in_var, width=10).pack(side="left")

            # 2. 淡出时长
            row2 = ttk.Frame(config_frame)
            row2.pack(fill="x", pady=8)
            ttk.Label(row2, text="淡出时长（秒）：", width=14, anchor="w").pack(side="left")
            fade_out_var = ttk.StringVar(value=self.config_values['C_fade_out'].get())
            ttk.Entry(row2, textvariable=fade_out_var, width=10).pack(side="left")

            # 提示信息
            ttk.Label(dialog, text="提示：\n淡入时长通常为0.1-0.5秒\n淡出时长通常为2.0-8.0秒", 
                     font=("Microsoft YaHei", 8)).pack()

            # 显示当前配置
            current_config = self.config_values['C'].get()
            ttk.Label(dialog, text=f"当前配置：{current_config}", 
                     font=("Microsoft YaHei", 9), bootstyle="info").pack(pady=5)

            # 按钮
            btn_frame = ttk.Frame(dialog)
            btn_frame.pack(pady=15)
            
            def save_config_c():
                fade_in = fade_in_var.get().strip()
                fade_out = fade_out_var.get().strip()

                # 验证
                try:
                    float(fade_in)
                    float(fade_out)
                except:
                    Messagebox.show_warning(
                        title="提示",
                        message="请输入有效的数字！",
                        parent=dialog
                    )
                    return
                
                if float(fade_in) < 0 or float(fade_out) < 0:
                    Messagebox.show_warning(
                        title="提示",
                        message="时长不能为负数！",
                        parent=dialog
                    )
                    return

                # 保存
                self.config_values['C_fade_in'] = ttk.StringVar(value=fade_in)
                self.config_values['C_fade_out'] = ttk.StringVar(value=fade_out)
                
                display_text = f"淡入{fade_in}s 淡出{fade_out}s"
                self.config_values['C'].set(display_text)
                
                dialog.destroy()
                Messagebox.show_info(
                    title="成功",  
                    message=f"配置已更新为：{display_text}",
                    parent=self.root  
                )
            
            ttk.Button(btn_frame, text="确定", command=save_config_c, bootstyle="success", width=10).pack(side="left", padx=10)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy, bootstyle="secondary", width=10).pack(side="left", padx=10)

        elif code == 'D':
            # ===== 功能D的专用配置对话框 =====
            dialog = ttk.Toplevel(self.root)
            dialog.title("前缀添加配置")
            dialog.geometry("700x580")

            # 弹窗居中
            root_x = self.root.winfo_rootx()
            root_y = self.root.winfo_rooty()
            root_width = self.root.winfo_width()
            root_height = self.root.winfo_height()
            x = root_x + (root_width - 700) // 2
            y = root_y + (root_height - 580) // 2
            dialog.geometry(f"700x580+{x}+{y}")
            
            dialog.transient(self.root)
            dialog.grab_set()
            
            ttk.Label(dialog, text="背景音乐配置", font=("Microsoft YaHei", 12, "bold")).pack(pady=10)

            # 配置框架
            config_frame = ttk.Frame(dialog)
            config_frame.pack(pady=10, padx=20, fill="x")

            # 1. 背景音乐总时长
            row1 = ttk.Frame(config_frame)
            row1.pack(fill="x", pady=5)
            ttk.Label(row1, text="BGM总时长（秒）：", width=16, anchor="w").pack(side="left")
            duration_var = ttk.StringVar(value=self.config_values['D_duration'].get())
            ttk.Entry(row1, textvariable=duration_var, width=10).pack(side="left")

            # 2. 背景音乐开始时间
            row2 = ttk.Frame(config_frame)
            row2.pack(fill="x", pady=5)
            ttk.Label(row2, text="BGM开始时间（秒）：", width=16, anchor="w").pack(side="left")
            start_var = ttk.StringVar(value=self.config_values['D_start'].get())
            ttk.Entry(row2, textvariable=start_var, width=10).pack(side="left")

            # 3. 语音位置
            row3 = ttk.Frame(config_frame)
            row3.pack(fill="x", pady=5)
            ttk.Label(row3, text="语音位置：", width=16, anchor="w").pack(side="left")
            position_var = ttk.StringVar(value=self.config_values['D_position'].get())
            position_combo = ttk.Combobox(row3, textvariable=position_var, 
                                          values=["start", "middle", "end", "自定义秒数"], 
                                          width=12)
            position_combo.pack(side="left", padx=5)

            # 自定义秒数输入（仅当选择"自定义秒数"时显示）
            row3b = ttk.Frame(config_frame)
            row3b.pack(fill="x", pady=5)
            ttk.Label(row3b, text="自定义秒数：", width=16, anchor="w").pack(side="left")
            custom_position_var = ttk.StringVar(value=self.config_values['D_custom_position'].get())
            ttk.Entry(row3b, textvariable=custom_position_var, width=10).pack(side="left")
            ttk.Label(row3b, text="（选择'自定义秒数'时生效）", font=("Microsoft YaHei", 8)).pack(side="left", padx=5)

            # 4. BGM音量
            row4 = ttk.Frame(config_frame)
            row4.pack(fill="x", pady=5)
            ttk.Label(row4, text="BGM音量：", width=16, anchor="w").pack(side="left")
            volume_var = ttk.StringVar(value=self.config_values['D_volume'].get())
            ttk.Entry(row4, textvariable=volume_var, width=10).pack(side="left")
            ttk.Label(row4, text="（0.1 ~ 1.0）", font=("Microsoft YaHei", 8)).pack(side="left", padx=5)

            # 显示当前配置
            current_config = self.config_values['D'].get()
            ttk.Label(dialog, text=f"当前配置：{current_config}", 
                     font=("Microsoft YaHei", 9), bootstyle="info").pack(pady=8)

            # 按钮
            btn_frame = ttk.Frame(dialog)
            btn_frame.pack(pady=15)
            
            def save_config_d():
                duration = duration_var.get().strip()
                start = start_var.get().strip()
                position = position_var.get()
                custom_position = custom_position_var.get().strip()
                volume = volume_var.get().strip()

                # 验证
                try:
                    float(duration)
                    float(start)
                    float(volume)
                except:
                    Messagebox.show_warning(
                        title="提示",
                        message="请输入有效的数字！",
                        parent=dialog
                    )
                    return
                
                if float(volume) < 0.1 or float(volume) > 1.0:
                    Messagebox.show_warning(
                        title="提示",
                        message="音量范围应在 0.1 ~ 1.0 之间！",
                        parent=dialog
                    )
                    return

                # 处理语音位置
                if position == "自定义秒数":
                    try:
                        float(custom_position)
                        final_position = custom_position
                    except:
                        Messagebox.show_warning(
                            title="提示",
                            message="自定义秒数请输入有效数字！",
                            parent=dialog
                        )
                        return
                else:
                    final_position = position

                # 保存
                self.config_values['D_duration'] = ttk.StringVar(value=duration)
                self.config_values['D_start'] = ttk.StringVar(value=start)
                self.config_values['D_position'] = ttk.StringVar(value=position)
                self.config_values['D_custom_position'] = ttk.StringVar(value=custom_position)
                self.config_values['D_volume'] = ttk.StringVar(value=volume)
                
                display_text = f"时长{duration}s 音量{volume} 位置{final_position}"
                self.config_values['D'].set(display_text)
                
                dialog.destroy()
                Messagebox.show_info(
                    title="成功",  
                    message=f"配置已更新为：{display_text}",
                    parent=self.root  
                )
            
            ttk.Button(btn_frame, text="确定", command=save_config_d, bootstyle="success", width=10).pack(side="left", padx=10)
            ttk.Button(btn_frame, text="取消", command=dialog.destroy, bootstyle="secondary", width=10).pack(side="left", padx=10)

    def run_search(self):
        if not self.data_loaded:
            Messagebox.show_error(
                title="错误",
                message="请先加载数据！",
                parent=self.root
            )
            return
        
        active_rules = {key: var.get() for key, var in self.rules.items()}
        
        try:
            gap = int(self.gap_value.get())
            if gap < 1:
                Messagebox.show_warning(
                    title="提示",
                    message="间隔值必须 ≥ 1",
                    parent=self.root
                )
                return
        except:
            Messagebox.show_warning(
                title="提示",
                message="请输入有效的间隔数字",
                parent=self.root
            )
            return
        
        active_count = sum(active_rules.values())
        if active_count == 0 and not self.fixed_positions:
            Messagebox.show_warning(
                title="提示",
                message="请至少选择一条规则或设置固定位置！",
                parent=self.root
            )
            return
        
        try:
            max_attempts = int(self.max_attempts.get())
        except:
            max_attempts = 50000
        
        self.btn_generate.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.searching = True

        # ===== 按钮触发弹窗新增的3行代码，开始新搜索时禁用保存按钮并清空之前的结果=====
        self.btn_save_result.config(state="disabled")  # 禁用"保存排布结果"按钮
        self.last_dance_list = None  # 清空之前保存的舞种列表
        self.last_seq = None  # 清空之前保存的序列数据
        # ===== 新增结束 =====

        self.result_text.delete("1.0", "end")
        self.result_text.insert("end", "正在搜索，请稍候...\n")
        enabled = [k for k, v in active_rules.items() if v]
        self.result_text.insert("end", f"已启用规则: {', '.join(enabled) if enabled else '无'}\n")
        self.result_text.insert("end", f"同舞种间隔值: {self.gap_value.get()}\n")
        self.result_text.insert("end", f"固定位置: {self.fixed_positions}\n")
        self.result_text.insert("end", f"最大搜索次数: {max_attempts}\n")
        self.result_text.insert("end", "-" * 50 + "\n")
        self.root.update()
        
        self._search_step(active_rules, 0, max_attempts)
    
    def _search_step(self, active_rules, attempt, max_attempts):
        if not self.searching:
            return
        
        if attempt >= max_attempts:
            # ... 失败处理 ...
            self.result_text.insert("end", "\n" + "=" * 50 + "\n")
            self.result_text.insert("end", f"❌ 搜索 {max_attempts} 次未找到合法解。\n")
            self.result_text.insert("end", "建议操作：\n")
            self.result_text.insert("end", "  1. 取消勾选 '节奏约束'（如果已启用）\n")
            self.result_text.insert("end", "  2. 将 '同舞种间隔值' 调低（如从6降到5或4）\n")
            self.result_text.insert("end", "  3. 增加 '最大搜索次数'（如 100000）\n")
            self.result_text.insert("end", "  4. 减少固定位置数量\n")
            self.progress_label.config(text=f"搜索失败，已尝试 {max_attempts} 次")
            self.btn_generate.config(state="normal")
            self.btn_stop.config(state="disabled")
            return
        
        if attempt % 500 == 0:
            self.progress_label.config(text=f"正在搜索... 已尝试 {attempt} 次")
            self.root.update()
        
        seq = self.generate_sequence(active_rules, max_attempts=1)
        
        # 增加将排布结果存储到表格：
        if seq is not None:
            # ===== 搜索成功 =====
            # 1. 保存结果
            self.last_attempt = attempt
            self.result_text.delete("1.0", "end")
            self.result_text.insert("end", f"✅ 找到合法解！(尝试 {attempt+1} 次)\n")
            self.result_text.insert("end", "-" * 38 + "\n")
            self.result_text.insert("end", f"{'序号':<4} {'大类':<8} {'舞种':<12} {'节奏':<4}\n")
            self.result_text.insert("end", "-" * 38 + "\n")
            for i, (cat, dance, rhythm) in enumerate(seq, 1):
                self.result_text.insert("end", f"{i:<4} {cat:<8} {dance:<12} {rhythm:<4}\n")

            # 2. 显示结果,舞种列输出
            self.result_text.insert("end", "\n" + "-" * 38 + "\n")
            self.result_text.insert("end", "【仅舞种列表】（可复制使用）\n")
            self.result_text.insert("end", "-" * 38 + "\n")
            dance_list = [dance for _, dance, _ in seq]
            self.result_text.insert("end", "、".join(dance_list) + "\n")
            self.result_text.insert("end", "-" * 38 + "\n")
            self.result_text.insert("end", f"共 {len(dance_list)} 首\n")
            
            # 弹窗询问是否保存到Excel
            # if messagebox.askyesno("保存结果", "是否将舞种排布结果保存到Excel文件？"):
            # 保存当前结果到实例变量，供"保存排布结果"按钮使用
            self.last_dance_list = dance_list
            self.last_seq = seq
            self.last_max_attempts = max_attempts

            # 3. 启用保存按钮
            self.btn_save_result.config(state="normal")  # 启用保存按钮

            # ===== 新增：校验并显示结果 =====
            ok, errors = self.check_constraints(seq, active_rules)
            self.result_text.insert("end", "\n" + "-" * 38 + "\n")
            if ok:
                self.result_text.insert("end", "✅ 所有启用的规则均已通过校验\n")
            else:
                self.result_text.insert("end", f"⚠️ 发现 {len(errors)} 个问题:\n")
                for e in errors[:10]:
                    self.result_text.insert("end", f"  - {e}\n")
            
            # 4. 更新进度标签和按钮状态
            self.progress_label.config(text=f"搜索成功！共尝试 {attempt+1} 次")
            self.btn_generate.config(state="normal")
            self.btn_stop.config(state="disabled")
            self.searching = False
            return  # 关键：成功时直接返回，不再继续搜索
        
        # 搜索失败：继续下一次搜索
        self.root.after(5, self._search_step, active_rules, attempt + 1, max_attempts)

    def save_result_to_excel(self):
        """保存排布结果到Excel（由按钮触发）"""
        if self.last_dance_list is None:
            Messagebox.show_warning(
                title="提示",
                message="没有可保存的排布结果，请先搜索！",
                parent=self.root
            )
            return
        
        try:
            # 获取当前脚本所在目录
            script_dir = os.path.dirname(os.path.abspath(__file__))
            file_path = os.path.join(script_dir, "舞曲排布_示例.xlsx")
            sheet_name = "排布结果"
            
            # 检查文件是否存在
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
            else:
                wb = openpyxl.Workbook()
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # 获取或创建"排布结果"工作表
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # 检测第一个空列
            col_idx = 1
            while ws.cell(row=1, column=col_idx).value is not None:
                col_idx += 1
            
            # 写入舞种列表
            for i, dance in enumerate(self.last_dance_list, 1):
                ws.cell(row=i, column=col_idx, value=dance)

            # ---------- 写入生成条件到"条件记录"工作表 ----------
            # 获取或创建"条件记录"工作表            
            if "条件记录" in wb.sheetnames:
                ws_conditions = wb["条件记录"]
            else:
                ws_conditions = wb.create_sheet("条件记录")
            
            # 获取列字母
            col_letter = chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"
            
            # 获取当前日期时间（月日时分）
            import datetime
            now = datetime.datetime.now()
            datetime_str = now.strftime("%m月%d日%H时%M分")
            
            # 构建条件信息
            fixed_pos_str = ""
            if self.fixed_positions:
                fixed_items = [f"{pos}={dance}" for pos, dance in sorted(self.fixed_positions.items())]
                fixed_pos_str = "，".join(fixed_items)
            else:
                fixed_pos_str = "无固定位置"

            # 规则信息
            # active_rules = [key for key, var in self.rules.items() if var.get()](AttributeError: 'list' object has no attribute 'get')
            active_rules = {key: var.get() for key, var in self.rules.items()}
            rules_str = "、".join([k for k, v in active_rules.items() if v]) if any(active_rules.values()) else "无规则"

            # 同舞种间隔值
            gap_str = f"间隔{self.gap_value.get()}首" if self.rules['同舞种间隔'].get() else ""

            # 最近的一次搜索次数
            attempts_str = f"搜索{self.last_max_attempts}次"

            # 功能配置信息
            func_configs = []
            if hasattr(self, 'config_values'):
                for code in ['A', 'B', 'C', 'D']:
                    if code in self.config_values:
                        func_configs.append(f"{code}:{self.config_values[code].get()}")
            func_str = "；".join(func_configs) if func_configs else "无功能配置"

            # 组合所有条件信息
            # condition_info = f"固定位置：{fixed_pos_str}；规则：{rules_str}；{gap_str}；{attempts_str}；功能配置：{func_str}"

            # 在"条件记录"工作表中查找第一个空行（从第2行开始）
            row_idx = 2
            while ws_conditions.cell(row=row_idx, column=1).value is not None:
                row_idx += 1   # 如果要跳行就改成2

            # 写入条件记录（拆分到各列）
            ws_conditions.cell(row=row_idx, column=1, value=datetime_str)       # A列：生成时间
            ws_conditions.cell(row=row_idx, column=2, value=self.last_attempt + 1)        # B列：尝试次数
            ws_conditions.cell(row=row_idx, column=3, value=col_letter )        # C列：结果列名
            ws_conditions.cell(row=row_idx, column=4, value=fixed_pos_str)      # D列：固定位置
            ws_conditions.cell(row=row_idx, column=5, value=gap_str)            # E列：舞种间隔
            ws_conditions.cell(row=row_idx, column=6, value=rules_str)          # F列：应用规则
            # ws_conditions.cell(row=row_idx, column=7, value=func_str)           # G列：功能配置

            # 保存
            wb.save(file_path)

            # 确保参数顺序正确
            Messagebox.show_info(
                title="保存成功",  # 第一个参数是标题
                message=f"✅ 已保存到：\n{file_path}\n\n"
                        f"舞种列表：在{sheet_name}工作表 {col_letter}列（共{len(self.last_dance_list)}首）\n\n"
                        f"生成条件：在条件记录工作表 第{row_idx}行",  # 第二个参数是消息
                parent=self.root  # 可选的父窗口
            )
                        
        except PermissionError:
            Messagebox.show_error(
                title="错误",
                message="文件正在使用中，请关闭Excel文件后重试！",
                parent=self.root
            )
        except Exception as e:
            Messagebox.show_error(
                title="错误",
                message=f"保存失败：{str(e)}",
                parent=self.root
            )

# ---------- 启动 ----------
if __name__ == "__main__":

    # 使用ttkbootstrap创建主窗口
    root = ttk.Window(themename="pulse-dark")
    
    # 设置窗口图标
    icon_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "icon.ico")
    if os.path.exists(icon_path):
        root.iconbitmap(icon_path)
    
    app = PlaylistApp(root)
    root.mainloop()